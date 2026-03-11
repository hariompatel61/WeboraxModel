"""
Excel Tracker: Records content details and video links for each pipeline run.
Maintains a content_log.xlsx file for tracking all generated content.
"""

import datetime
import os

from app_config import Config

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class ExcelTracker:
    """Track pipeline runs in an Excel spreadsheet."""

    HEADERS = [
        "Sr. No",
        "Run ID",
        "Date & Time",
        "Topic Title",
        "Category",
        "Premise",
        "Script Hook",
        "Characters",
        "Video Path",
        "YouTube Link",
        "Status",
    ]

    def __init__(self, filepath=None):
        self.filepath = filepath or os.path.join(Config.OUTPUT_DIR, "content_log.xlsx")
        self._ensure_workbook()

    def _ensure_workbook(self):
        """Create the workbook with headers if it does not exist."""
        if not OPENPYXL_AVAILABLE:
            print("   Warning: openpyxl not installed. Cannot create/log to Excel.")
            print("   Run: pip install openpyxl")
            return

        if os.path.exists(self.filepath):
            return

        os.makedirs(os.path.dirname(self.filepath), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "Content Log"
        ws.append(self.HEADERS)

        header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_num in range(1, len(self.HEADERS) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        widths = [8, 22, 20, 25, 18, 40, 30, 25, 40, 45, 12]
        for col_num, width in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

        ws.freeze_panes = "A2"
        wb.save(self.filepath)
        print(f"   Created new Excel log: {self.filepath}")

    def log_run(
        self,
        topic_data=None,
        script_data=None,
        video_path=None,
        yt_link=None,
        upload_result=None,
        status="Completed",
        run_id=None,
        topic=None,
        script=None,
    ):
        """Log a pipeline run to the tracker file."""
        if not OPENPYXL_AVAILABLE:
            print("   Warning: openpyxl not installed. Cannot log to Excel.")
            print("   Run: pip install openpyxl")
            return None

        self._ensure_workbook()
        wb = load_workbook(self.filepath)
        ws = wb.active
        next_row = ws.max_row + 1

        topic_data = topic_data or topic or {}
        script_data = script_data or script or {}
        upload_result = upload_result or {}
        if not yt_link:
            yt_link = upload_result.get("url", "")

        characters = topic_data.get("characters", [])
        if isinstance(characters, (list, tuple)):
            characters = ", ".join(str(char) for char in characters)

        run_id = run_id or f"run_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
        row_data = [
            next_row - 1,
            run_id,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            topic_data.get("title", topic_data.get("topic", "")),
            topic_data.get("category", ""),
            topic_data.get("premise", ""),
            script_data.get("hook", ""),
            characters,
            video_path or "",
            yt_link or "",
            status,
        ]

        data_align = Alignment(vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        status_colors = {
            "Completed": "C6EFCE",
            "Failed": "FFC7CE",
            "Partial": "FFEB9C",
        }

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_num, value=value)
            cell.alignment = data_align
            cell.border = thin_border
            if col_num == len(self.HEADERS):
                fill = status_colors.get(status, "FFFFFF")
                cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

        wb.save(self.filepath)
        print(f"   Logged to Excel: {self.filepath}")
        return self.filepath
